import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
wb = openpyxl.Workbook()
HF = PatternFill("solid", fgColor="1F4E79")
HN = Font(bold=True, color="FFFFFF", size=11)
AF = PatternFill("solid", fgColor="D6E4F0")
BD = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))

def ss(ws, headers, rows, cw=None):
    for col,h in enumerate(headers,1):
        c=ws.cell(row=1,column=col,value=h)
        c.font=HN; c.fill=HF
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border=BD
    ws.row_dimensions[1].height=30
    for r,row in enumerate(rows,2):
        fill=AF if r%2==0 else PatternFill()
        for col,val in enumerate(row,1):
            c=ws.cell(row=r,column=col,value=val)
            c.fill=fill; c.alignment=Alignment(vertical="center",wrap_text=True); c.border=BD
        ws.row_dimensions[r].height=20
    if cw:
        for i,w in enumerate(cw,1): ws.column_dimensions[get_column_letter(i)].width=w
